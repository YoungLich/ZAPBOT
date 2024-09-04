from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
import openpyxl
from openpyxl import Workbook
from time import sleep
from pathlib import Path
import sys
import pyautogui as pa
import time

# Função para formatar o número do processo
def formatar_processo(proc):
    s = ''.join(filter(str.isdigit, proc))
    if len(s) < 20:
        s = s.zfill(20)
    return f"{s[:7]}-{s[7:9]}.{s[9:13]}.{s[13:14]}.{s[14:16]}.{s[16:20]}"

# Função para ler os processos a serem excluídos do arquivo Excel
def ler_arquivos_exclusao(arquivo_exclusao):
    processos = []
    try:
        workbook = openpyxl.load_workbook(arquivo_exclusao)
        planilha = workbook.active
        for linha in planilha.iter_rows(min_row=2, values_only=True):
            processo = formatar_processo(linha[0])
            processos.append(processo)
    except FileNotFoundError:
        print(f"Arquivo '{arquivo_exclusao}' não encontrado.")
        sys.exit(1)
    return processos

def execute_site_automation(username, password, save_directory):
    arquivo_excel = Path(__file__).parent / 'Excluir_Push.xlsx'
    try:
        workbook = openpyxl.load_workbook(arquivo_excel)
        ler_planilha = workbook['plan1']
    except FileNotFoundError:
        print(f"Arquivo '{arquivo_excel}' não encontrado. Verifique o caminho do arquivo.")
        sys.exit(1)

# Função para escrever o resultado da exclusão no arquivo Excel
def escrever_arquivo(processo, msg):
    arquivo_resultado = Path(__file__).parent / 'Excluidos.xlsx'
    try:
        if arquivo_resultado.exists():
            wb = openpyxl.load_workbook(arquivo_resultado)
        else:
            wb = Workbook()
            wb.active.title = "Excluídos"
            wb.active.append(["Processo", "Mensagem"])

        planilha = wb.active
        planilha.append([processo, msg])
        wb.save(arquivo_resultado)
        print(f"{processo}: {msg}")
    except Exception as e:
        print(f"Erro ao escrever no arquivo: {e}")

# Função principal para excluir processos
def excluir_processos(username, password, arquivo_exclusao):
    # Configurações do navegador
    options = webdriver.ChromeOptions()
    options.add_argument('--start-maximized')
    options.add_argument('--disable-infobars')
    options.add_argument('--disable-extensions')
    options.page_load_strategy = 'normal'
    options.add_experimental_option("detach", True)

    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(options=options, service=service)

    # Abrindo o site e fazendo login
    url_pje = "https://pje1g.trf3.jus.br/pje/login.seam"
    driver.get(url_pje)

    # Mudar para o iframe correto
    frame_correto = driver.find_element(By.XPATH, '//*[@id="ssoFrame"]')
    driver.switch_to.frame(frame_correto)

    # Preencher username e senha
    driver.find_element(By.ID, 'username').send_keys(username)
    driver.find_element(By.ID, 'password').send_keys(password)
    sleep(2)
    driver.find_element(By.ID, 'kc-login').click()

    # Verificação e navegação usando PyAutoGUI
    # Pausa entre as ações do PyAutoGUI
    pa.PAUSE = 1

    # Navegação manual com PyAutoGUI (ajuste as coordenadas conforme necessário)
    print("Iniciando navegação com PyAutoGUI...")
    time.sleep(5)  # Aguarda um tempo para carregar a página após o login

    pa.click(x=31, y=112)  # Primeiro clique após o login
    pa.click(x=115, y=225)  # Segundo clique para navegar
    pa.click(x=77, y=460)  # Terceiro clique
    pa.click(x=52, y=264)  # Quarto clique
    pa.click(x=212, y=308)  # Quinto clique para chegar na página do Push

    time.sleep(3)  # Espera para garantir que a página carregou completamente

    processos_para_excluir = ler_arquivos_exclusao(arquivo_exclusao)

    for processo in processos_para_excluir:
        # Navegar para o input de processos com PyAutoGUI
        pa.click(x=138, y=605)  # Clique no input de processos
        pa.typewrite(processo)  # Digita o número do processo
        pa.press('ENTER')  # Pressiona Enter para buscar o processo

        time.sleep(2)  # Espera o processo ser carregado

        # Clique para excluir o processo
        pa.click(x=138, y=605)  # Posiciona no botão de excluir
        pa.press('ENTER')  # Confirma a exclusão

        # Verificar se a exclusão foi bem-sucedida
        time.sleep(2)  # Espera a confirmação

        try:
            # Se a operação for concluída com sucesso, escreve no arquivo
            escrever_arquivo(processo, "Excluído com sucesso")
        except Exception as e:
            escrever_arquivo(processo, f'Erro ao excluir: {e}')

    driver.quit()

# Função para verificar se as credenciais foram salvas e iniciar o processo de exclusão
def iniciar_exclusao():
    username = input("Digite o username: ")
    password = input("Digite a senha: ")
    arquivo_exclusao = 'arquivo_exclusao.xlsx'

    # Aqui, o programa continua e executa a exclusão
    excluir_processos(username, password, arquivo_exclusao)

# Verificação e execução principal
if __name__ == "__main__":
    print("Aguardando verificação de credenciais...")
    iniciar_exclusao()
