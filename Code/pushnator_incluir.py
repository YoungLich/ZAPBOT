from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import Workbook
import openpyxl
from pathlib import Path
import json
import sys

def save_login_data(username, password):
    data = {"username": username, "password": password}
    with open("login_data.json", "w") as file:
        json.dump(data, file)

def load_login_data():
    try:
        with open("login_data.json", "r") as file:
            data = json.load(file)
            return data.get("username", ""), data.get("password", "")
    except FileNotFoundError:
        return "", ""

def execute_site_automation(username, password, file_path, save_directory):
    arquivo_excel = Path(file_path)
    try:
        workbook = openpyxl.load_workbook(arquivo_excel)
        ler_planilha = workbook.active
    except FileNotFoundError:
        print(f"Arquivo '{arquivo_excel}' não encontrado. Verifique o caminho do arquivo.")
        sys.exit(1)

    nova_planilha = Workbook()
    planilha_sheet = nova_planilha.active
    planilha_sheet.title = "Resultados"
    planilha_sheet.append(["Processo", "Status"])

    options = webdriver.ChromeOptions()
    options.add_argument('--start-maximized')
    options.add_argument('--disable-infobars')
    options.add_argument('--disable-extensions')
    options.page_load_strategy = 'normal'

    service = Service(executable_path=ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=options)

    try:
        url = "https://pje1g.trf3.jus.br/pje/Push/listView.seam"
        driver.get(url)

        frame_correta = driver.find_element(By.XPATH, '//*[@id="ssoFrame"]')
        driver.switch_to.frame(frame_correta)

        driver.find_element(By.ID, 'username').send_keys(username)
        driver.find_element(By.ID, 'password').send_keys(password)
        driver.find_element(By.ID, 'kc-login').click()

        WebDriverWait(driver, 10).until(EC.url_to_be(url))

        for linha in ler_planilha.iter_rows(min_row=1):
            processo = linha[0].value
            if processo:
                seletor_processo = driver.find_element(By.XPATH, '//*[@id="j_id148:inputNumeroProcesso-inputNumeroProcessoDecoration:inputNumeroProcesso-inputNumeroProcesso"]')
                seletor_processo.send_keys(processo)
                seletor_processo.send_keys(Keys.RETURN)

                wait = WebDriverWait(driver, 10)

                try:
                    resultado = wait.until(EC.presence_of_element_located((By.XPATH, f'//*[contains(text(), "{processo}")]')))
                    print(f'Processo {processo} já está no site.')
                    planilha_sheet.append([processo, "Já está no site"])
                except:
                    print(f'Processo {processo} não encontrado no site.')
                    planilha_sheet.append([processo, "Não encontrado no site"])

                    try:
                        botao_incluir = driver.find_element(By.XPATH, '//*[@id="j_id148:btnIncluirAcompanhamento"]')
                        botao_incluir.click()

                        input_processo = wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="j_id148:inputNumeroProcesso-inputNumeroProcessoDecoration:inputNumeroProcesso-inputNumeroProcesso"]')))
                        input_processo.send_keys(processo)

                        print(f'Processo {processo} foi incluído no site.')
                        planilha_sheet.append([processo, "Incluído no site"])
                    except Exception as e:
                        print(f'Erro ao incluir o processo {processo}: {e}')
                        planilha_sheet.append([processo, f'Erro ao incluir: {e}'])

        resultado_path = Path(save_directory) / 'Resultados_Inclusao_Processos.xlsx'
        nova_planilha.save(resultado_path)
        print(f"Planilha de resultados salva com sucesso em {resultado_path}.")
    finally:
        driver.quit()
