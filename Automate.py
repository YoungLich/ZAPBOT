import tkinter as tk
from tkinter import filedialog, messagebox
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
import openpyxl
from openpyxl import Workbook
from urllib.parse import quote
from time import sleep
from pathlib import Path
import sys
import time

# Função para executar o script Selenium no site
def execute_site_automation(url, save_directory):
    arquivo_excel = Path(__file__).parent / 'Inserir_Push.xlsx'
    try:
        workbook = openpyxl.load_workbook(arquivo_excel)
        ler_planilha = workbook['plan1']
    except FileNotFoundError:
        print(f"Arquivo '{arquivo_excel}' não encontrado. Verifique o caminho do arquivo.")
        sys.exit(1)

    nova_planilha = Workbook()
    planilha_sheet = nova_planilha.active
    planilha_sheet.title = "Resultados"
    planilha_sheet.append(["Processo", "Status"])

    options = webdriver.ChromeOptions()
    options.add_argument('--start-maximized')
    options.add_argument('--disable-infobars')
    options.add_argument('--disable-extensions')
    options.page_load_strategy = 'normal'
    options.add_experimental_option("detach", True)

    service = Service(executable_path=ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service)

    driver.get(url)

    frame_correta = driver.find_element(By.XPATH, '//*[@id="ssoFrame"]')
    driver.switch_to.frame(frame_correta)

    login = driver.find_element(By.ID, 'username').send_keys('52312819805')
    senha = driver.find_element(By.ID, 'password').send_keys('Eg@5044326')
    sleep(2)
    driver.find_element(By.ID, 'kc-login').click()

    try:
        WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.XPATH, 'elemento_xpath_da_nova_pagina')))
        print("Login bem-sucedido e redirecionado para a nova página.")
    except Exception as e:
        print(f"Ocorreu um erro durante o login ou o redirecionamento: {e}")
        sleep(10)

    nova_url = "https://pje1g.trf3.jus.br/pje/Push/listView.seam"
    driver.get(nova_url)
    print(f"Redirecionado para {nova_url}")

    for linha in ler_planilha.iter_rows(min_row=1):
        processo = linha[0].value
        if processo:
            seletor_processo = driver.find_element(By.XPATH, '//*[@id="j_id148:inputNumeroProcesso-inputNumeroProcessoDecoration:inputNumeroProcesso-inputNumeroProcesso"]')
            seletor_processo.send_keys(processo)
            seletor_processo.send_keys(Keys.RETURN)

            wait = WebDriverWait(driver, 5)

            try:
                resultado = wait.until(EC.presence_of_element_located((By.XPATH, f'//[contains(text(), "{processo}")]')))
                sleep(10)
                print(f'Processo {processo} já está no site.')
                planilha_sheet.append([processo, "Já está no site"])
            except:
                print(f'Processo {processo} não encontrado no site.')
                planilha_sheet.append([processo, "Não encontrado no site"])

                try:
                    botao_incluir = driver.find_element(By.XPATH, '//*[@id="j_id148:btnIncluirAcompanhamento"]')
                    botao_incluir.click()

                    input_processo = wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="j_id148:inputNumeroProcesso-inputNumeroProcessoDecoration:inputNumeroProcesso-inputNumeroProcesso"]')))
                    input_processo.send_keys(processo)

                    print(f'Processo {processo} foi incluído no site.')
                    planilha_sheet.append([processo, "Incluído no site"])
                except Exception as e:
                    print(f'Erro ao incluir o processo {processo}: {e}')
                    planilha_sheet.append([processo, f'Erro ao incluir: {e}'])

    try:
        resultado_path = Path(save_directory) / 'Resultados_Inclusao_Processos.xlsx'
        nova_planilha.save(resultado_path)
        print(f"Planilha de resultados salva com sucesso em {resultado_path}.")
    except Exception as e:
        print(f"Erro ao salvar a planilha de resultados: {e}")

    driver.quit()

# Função para enviar mensagens no WhatsApp Web
def send_whatsapp_messages():
    driver = webdriver.Chrome()
    driver.get('https://web.whatsapp.com/')
    time.sleep(15)

    workbook = openpyxl.load_workbook('clientes.xlsx')
    pagina_clientes = workbook['Plan1']

    for linha in pagina_clientes.iter_rows(min_row=2):
        nome = linha[0].value
        telefone = linha[1].value
        mensagem = f'Olá {nome}, é apenas um teste'
        link_mens_whats = f'https://web.whatsapp.com/send?phone={telefone}&text={quote(mensagem)}'

        driver.get(link_mens_whats)
        time.sleep(10)

        try:
            seta = WebDriverWait(driver, 30).until(
                EC.element_to_be_clickable((By.XPATH, '//*[@id="main"]/footer/div[1]/div/span[2]/div/div[2]/div[2]/button/span'))
            )
            seta.click()
            time.sleep(5)
        except Exception as e:
            print(f'Não foi possível enviar a mensagem para {nome}: {str(e)}')
            with open('erros.csv', 'a', newline='', encoding='UTF-8') as arquivo:
                arquivo.write(f'{nome}, {telefone}\n')

    driver.quit()

# Função para abrir a janela de automação do site
def open_site_window():
    site_window = tk.Toplevel(root)
    site_window.title("Automação no Site")
    site_window.geometry("500x300")
    site_window.configure(bg="#F0F0F0")

    url_label = tk.Label(site_window, text="URL do Site:", bg="#F0F0F0")
    url_label.pack(pady=5)
    url_entry = tk.Entry(site_window, width=50)
    url_entry.pack(pady=5)

    def select_directory():
        directory = filedialog.askdirectory(title="Selecione o Diretório para Salvar a Planilha")
        if directory:
            directory_var.set(directory)
            directory_display.config(state="normal")
            directory_display.delete(1.0, tk.END)
            directory_display.insert(tk.END, directory)
            directory_display.config(state="disabled")

    directory_var = tk.StringVar()
    directory_button = tk.Button(site_window, text="Selecionar Diretório", command=select_directory)
    directory_button.pack(pady=5)

    directory_display = tk.Text(site_window, height=2, width=50, state="disabled", wrap="word")
    directory_display.pack(pady=5)

    execute_button = tk.Button(
        site_window,
        text="Executar Automação",
        command=lambda: execute_site_automation(url_entry.get(), directory_var.get()),
        bg="#4CAF50",
        fg="white",
        font=("Arial", 12),
        width=20,
        height=2
    )
    execute_button.pack(pady=20)

# Função para abrir a janela de envio de WhatsApp
def open_whatsapp_window():
    whatsapp_window = tk.Toplevel(root)
    whatsapp_window.title("Envio de Mensagens no WhatsApp")
    whatsapp_window.geometry("600x400")
    whatsapp_window.configure(bg="#F0F0F0")

    # Rótulo e campo de entrada para o número de telefone
    phone_label = tk.Label(whatsapp_window, text="Número de Telefone:", bg="#F0F0F0")
    phone_label.pack(pady=5)
    phone_entry = tk.Entry(whatsapp_window, width=30)
    phone_entry.pack(pady=5)

    # Rótulo e caixa de texto para a mensagem
    message_label = tk.Label(whatsapp_window, text="Mensagem:", bg="#F0F0F0")
    message_label.pack(pady=5)
    message_text = tk.Text(whatsapp_window, height=5, width=50)
    message_text.pack(pady=5)

    # Rótulo e campo de entrada para o horário de envio
    time_label = tk.Label(whatsapp_window, text="Horário de Envio (HH:MM):", bg="#F0F0F0")
    time_label.pack(pady=5)
    time_entry = tk.Entry(whatsapp_window, width=10)
    time_entry.pack(pady=5)

    # Função para enviar a mensagem no horário especificado
    def send_scheduled_message():
        phone = phone_entry.get()
        message = message_text.get("1.0", tk.END).strip()
        send_time = time_entry.get()

        # Verificar o horário atual e aguardar até o horário especificado
        current_time = time.strftime('%H:%M')
        while current_time != send_time:
            current_time = time.strftime('%H:%M')
            time.sleep(30)  # Checa a cada 30 segundos

        # Após o horário especificado, enviar a mensagem
        driver = webdriver.Chrome()
        driver.get(f'https://web.whatsapp.com/send?phone={phone}&text={quote(message)}')
        time.sleep(15)

        try:
            seta = WebDriverWait(driver, 30).until(
                EC.element_to_be_clickable((By.XPATH, '//*[@id="main"]/footer/div[1]/div/span[2]/div/div[2]/div[2]/button/span'))
            )
            seta.click()
            time.sleep(5)
        except Exception as e:
            print(f'Não foi possível enviar a mensagem: {str(e)}')
        finally:
            driver.quit()

    # Botão para enviar a mensagem no horário especificado
    send_button = tk.Button(
        whatsapp_window,
        text="Agendar Envio",
        command=send_scheduled_message,
        bg="#25D366",
        fg="white",
        font=("Arial", 12),
        width=20,
        height=2
    )
    send_button.pack(pady=20)


    execute_button = tk.Button(
        whatsapp_window,
        text="Enviar Mensagens",
        command=send_whatsapp_messages,
        bg="#25D366",
        fg="white",
        font=("Arial", 12),
        width=20,
        height=2
    )
    execute_button.pack(pady=60)

# Cria a janela principal
root = tk.Tk()
root.title("Automação Completa")
root.geometry("600x400")
root.configure(bg="#191970")

site_button = tk.Button(
    root,
    text="Pushnator",
    command=open_site_window,
    bg="black",
    fg="white",
    font=("Arial", 15, "bold"),
    width=15,
    height=2
)
site_button.grid(row=4, column=1, padx=50, pady=10)


whatsapp_button = tk.Button(
    root,
    text="ZapBot",
    command=open_whatsapp_window,
    bg="#3CB371",
    fg="black",
    font=("Arial", 15, "bold"),
    width=15,
    height=2
)
whatsapp_button.grid(row=0, column=0, padx=50, pady=10)

root.mainloop()